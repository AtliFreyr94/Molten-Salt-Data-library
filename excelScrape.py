import openpyxl
from openpyxl.utils import get_column_letter
import numpy

def excelScrape(fileName):
    dataSetList = []
    wb = openpyxl.load_workbook(fileName)
    #For each sheet we get 1 dataset, we want to loop through all the sheets
    for sheetName in wb.sheetnames:
        sheet = wb[sheetName]
        dataSet = {}

        #This only works if propertynames are stored in column A
        #This loops through all column A and grabs the values stored in cells to the right
        for cell in tuple(sheet.columns)[0]:
            propertyName = cell.value
            if propertyName == None:
                continue
            row = cell.row
            col = cell.column
            newCell = 'B' + str(row)
            dataList = []
            i = 3
            #This loop ensures that we keep grabbing additional cells as long as there is something in them
            while sheet[newCell].value != None:
                dataList.append(sheet[newCell].value)
                newCol = get_column_letter(i)
                newCell = newCol + str(row)
                i += 1
            #Convert all the data into numpy arrays, does't work on laptop, uncomment it later
            #dataList = numpy.array(dataList)

            #Add into dataSet
            dataSet[propertyName] = dataList
        dataSetList.append(dataSet)
    return dataSetList

#Example call
excelScrape('Cantor-1973.xlsx')
