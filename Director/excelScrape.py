#excelScrape.py -- File that scrapes the excel file indicated by other files and creates the data in a asdf ready format

#Feature to be added: Ability to scrape only one sheet specified by user

import openpyxl
from openpyxl.utils import get_column_letter
import numpy

#Main function, scrapes the entire excel file. Only works if the structure is as follows:
# Property names in the A column and values in the columns to the right of the property name,
# There can not be any gaps in data otherwise the scraper stops as the while loop only checks if the next column is not empty

#TODO: Refactor code so that scraping an excel sheet is it's seperate function
def excelScrape(fileName):
    dataSetList = []
    wb = openpyxl.load_workbook(fileName)
    #For each sheet we get 1 dataset, we want to loop through all the sheets
    for sheetName in wb.sheetnames:
        sheet = wb[sheetName]
        dataSet = {}

        #This only works if propertynames are stored in column A
        #This loops through all column A and grabs the values stored in cells to the right
        for cell in tuple(sheet.columns)[0]:
            propertyName = cell.value
            if propertyName == None:
                continue
            row = cell.row
            newCell = 'B' + str(row)
            dataList = []
            i = 3
            #This loop ensures that we keep grabbing additional cells as long as there is something in them
            while sheet[newCell].value != None:
                dataList.append(sheet[newCell].value)
                newCol = get_column_letter(i)
                newCell = newCol + str(row)
                i += 1

            #Convert the data to numpy array only if they are datasets containing floating point numbers
            if len(dataList) > 1 and (type(dataList[0]) == float or type([dataList[0]]) == int):
                dataList = numpy.array(dataList)

            #Add into dataSet
            dataSet[propertyName] = dataList
        dataSetList.append(dataSet)
    return dataSetList

def excelScrapeOneSheet(fileName):
    wb = openpyxl.load_workbook(fileName)
    dataSet = {}
    print (wb.sheetnames)
    sheetName = input('Enter the name of the excelsheet to scrape: ')
    sheet = wb[sheetName]
    for cell in tuple(sheet.columns)[0]:
        propertyName = cell.value
        if propertyName == None:
            continue
        row = cell.row
        newCell = 'B' + str(row)
        dataList = []
        i = 3
        #This loop ensures that we keep grabbing additional cells as long as there is something in them
        while sheet[newCell].value != None:
            dataList.append(sheet[newCell].value)
            newCol = get_column_letter(i)
            newCell = newCol + str(row)
            i += 1

        #Convert the data to numpy array only if they are datasets containing floating point numbers
        if len(dataList) > 1 and (type(dataList[0]) == float or type([dataList[0]]) == int):
            dataList = numpy.array(dataList)

        #Add into dataSet
        dataSet[propertyName] = dataList
    return dataSet
